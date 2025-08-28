import os
from openpyxl import Workbook
import pdfplumber
import re

directory = './PDFS/MetLife'
files = os.listdir(directory)
files_quantity = len(files)

if files_quantity == 0:
    raise Exception('Files não encontrados')

wb = Workbook()
ws = wb.active
ws.title = 'Pdf imports'

ws['A1'] = 'Nº Cirurgia'
ws['B1'] = 'Nome Cirurgia'
ws['C1'] = 'Porcentagem'
ws['D1'] = 'Status'

last_empty_line = 1

pattern = re.compile(r'(\d+)\s+(.+?)\s+(\d+%)')

for file in files:
    with pdfplumber.open(os.path.join(directory, file)) as pdf:
        for page in pdf.pages:
            pdf_text = page.extract_text()
            if pdf_text:
                for match in pattern.findall(pdf_text):
                    numero, nome, porcentagem = match
                    ws[f"A{last_empty_line}"] = numero
                    ws[f"B{last_empty_line}"] = nome.strip()
                    ws[f"C{last_empty_line}"] = porcentagem
                    ws[f"D{last_empty_line}"] = 'OK'
                    last_empty_line += 1

wb.save('cirurgias.xlsx')